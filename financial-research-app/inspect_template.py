#!/usr/bin/env python3
"""Inspect Excel template structure"""

import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent))

from openpyxl import load_workbook

# Load the most recent template
template_path = 'data/uploads/20251210_180355_Format-NEW.xlsx'

try:
    wb = load_workbook(template_path)
    print("=" * 80)
    print(f"Template: {template_path}")
    print(f"Sheets: {wb.sheetnames}")
    print("=" * 80)
    
    # Check Sheet 3 (or last sheet)
    if len(wb.worksheets) >= 3:
        ws = wb.worksheets[2]
        sheet_name = wb.sheetnames[2]
    else:
        ws = wb.active
        sheet_name = wb.active.title
    
    print(f"\nSheet: {sheet_name}")
    print("-" * 80)
    
    # Print first 25 rows, first 15 columns
    for row_idx in range(1, 26):
        row_data = []
        for col_idx in range(1, 16):
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            if cell_value:
                row_data.append(f"{col_idx}:{cell_value}")
        if row_data:
            print(f"Row {row_idx:2d}: {' | '.join(row_data)}")
    
    print("=" * 80)
    
except Exception as e:
    print(f"Error: {e}")
    import traceback
    traceback.print_exc()
