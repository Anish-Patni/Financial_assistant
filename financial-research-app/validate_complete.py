#!/usr/bin/env python3
"""
Complete End-to-End Validation
Validates entire system with user's Excel input format
"""

import sys
from pathlib import Path
sys.path.insert(0, str(Path(__file__).parent))

from core.excel_generator import ExcelGenerator
from core.research_storage import ResearchStorage
from core.data_extractor import FinancialDataExtractor
from config import settings
from config.logging_config import setup_logging
import openpyxl

setup_logging(log_level='INFO')

def validate_excel_input(excel_path: str):
    """Validate user's Excel input file"""
    print("\n" + "="*80)
    print("  VALIDATING USER EXCEL INPUT")
    print("="*80 + "\n")
    
    try:
        wb = openpyxl.load_workbook(excel_path)
        sheet = wb.active
        
        # Extract company names from first column
        companies = []
        row = 2  # Start from row 2 (after header)
        while True:
            cell_value = sheet.cell(row=row, column=1).value
            if cell_value is None or cell_value == '':
                break
            companies.append(cell_value)
            row += 1
        
        print(f"✓ Excel file loaded successfully")
        print(f"✓ Found {len(companies)} companies:")
        for i, company in enumerate(companies, 1):
            print(f"   {i}. {company}")
        
        # Validate column structure
        headers = []
        for col in range(1, 20):
            header = sheet.cell(row=1, column=col).value
            if header:
                headers.append(header)
        
        print(f"\n✓ Found {len(headers)} data columns")
        print(f"  Headers: {', '.join(headers[:5])}...")
        
        return {
            'success': True,
            'companies': companies,
            'headers': headers,
            'row_count': len(companies)
        }
        
    except Exception as e:
        print(f"✗ Error validating Excel: {e}")
        return {'success': False, 'error': str(e)}

def validate_research_data():
    """Validate existing research data"""
    print("\n" + "="*80)
    print("  VALIDATING RESEARCH DATA")
    print("="*80 + "\n")
    
    storage = ResearchStorage(settings.RESEARCH_RESULTS_DIR)
    all_research = storage.get_all_research()
    summary = storage.get_research_summary()
    
    print(f"✓ Storage system operational")
    print(f"✓ Total research results: {summary['total_research_count']}")
    print(f"✓ Companies researched: {', '.join(summary['companies'])}")
    print(f"✓ Total indicators extracted: {summary['total_extractions']}")
    print(f"✓ Average confidence: {summary['average_confidence']*100:.1f}%")
    
    return {
        'success': True,
        'research_count': summary['total_research_count'],
        'companies': summary['companies'],
        'confidence': summary['average_confidence']
    }

def validate_excel_generation():
    """Validate Excel generation from research data"""
    print("\n" + "="*80)
    print("  VALIDATING EXCEL GENERATION")
    print("="*80 + "\n")
    
    try:
        storage = ResearchStorage(settings.RESEARCH_RESULTS_DIR)
        all_results = storage.get_all_research()
        
        if not all_results:
            print("✗ No research data available for Excel generation")
            return {'success': False, 'error': 'No data'}
        
        generator = ExcelGenerator()
        output_path = Path('data/output/validation_output.xlsx')
        
        success = generator.generate_excel(all_results, output_path)
        
        if success and output_path.exists():
            file_size = output_path.stat().st_size
            print(f"✓ Excel file generated successfully")
            print(f"✓ Output file: {output_path}")
            print(f"✓ File size: {file_size:,} bytes")
            
            # Validate output structure
            wb = openpyxl.load_workbook(output_path)
            print(f"✓ Sheets created: {', '.join(wb.sheetnames)}")
            
            # Check summary sheet
            summary_sheet = wb['Summary']
            data_rows = 0
            for row in range(4, 100):
                if summary_sheet.cell(row=row, column=1).value:
                    data_rows += 1
                else:
                    break
            
            print(f"✓ Summary sheet has {data_rows} data rows")
            
            return {
                'success': True,
                'file_path': str(output_path),
                'file_size': file_size,
                'sheets': len(wb.sheetnames),
                'data_rows': data_rows
            }
        else:
            print("✗ Excel generation failed")
            return {'success': False, 'error': 'Generation failed'}
            
    except Exception as e:
        print(f"✗ Error: {e}")
        import traceback
        traceback.print_exc()
        return {'success': False, 'error': str(e)}

def validate_system_components():
    """Validate all system components"""
    print("\n" + "="*80)
    print("  VALIDATING SYSTEM COMPONENTS")
    print("="*80 + "\n")
    
    results = {
        'config': False,
        'storage': False,
        'extractor': False,
        'generator': False
    }
    
    # Test configuration
    try:
        assert settings.RESEARCH_RESULTS_DIR is not None
        assert Path(settings.RESEARCH_RESULTS_DIR).exists()
        results['config'] = True
        print("✓ Configuration system working")
    except Exception as e:
        print(f"✗ Configuration error: {e}")
    
    # Test storage
    try:
        storage = ResearchStorage(settings.RESEARCH_RESULTS_DIR)
        summary = storage.get_research_summary()
        results['storage'] = True
        print(f"✓ Storage system working ({summary['total_research_count']} results)")
    except Exception as e:
        print(f"✗ Storage error: {e}")
    
    # Test extractor
    try:
        extractor = FinancialDataExtractor()
        test_text = "Revenue was ₹5000 crores with EBITDA of ₹1200 crores"
        indicators = extractor.extract_indicators(test_text, "Q1", 2024)
        results['extractor'] = True
        print(f"✓ Data extractor working ({len(indicators)} patterns)")
    except Exception as e:
        print(f"✗ Extractor error: {e}")
    
    # Test generator
    try:
        generator = ExcelGenerator()
        results['generator'] = True
        print("✓ Excel generator initialized")
    except Exception as e:
        print(f"✗ Generator error: {e}")
    
    all_passed = all(results.values())
    return {'success': all_passed, 'components': results}

def run_complete_validation(excel_path: str = None):
    """Run complete end-to-end validation"""
    print("\n" + "="*80)
    print("  COMPLETE SYSTEM VALIDATION")
    print("  AI Financial Research Application")
    print("="*80)
    
    validation_results = []
    
    # 1. Validate system components
    result = validate_system_components()
    validation_results.append(('Components', result['success']))
    
    # 2. Validate user Excel input (if provided)
    if excel_path and Path(excel_path).exists():
        result = validate_excel_input(excel_path)
        validation_results.append(('Excel Input', result['success']))
    
    # 3. Validate research data
    result = validate_research_data()
    validation_results.append(('Research Data', result['success']))
    
    # 4. Validate Excel generation
    result = validate_excel_generation()
    validation_results.append(('Excel Output', result['success']))
    
    # Print summary
    print("\n" + "="*80)
    print("  VALIDATION SUMMARY")
    print("="*80 + "\n")
    
    all_passed = True
    for test_name, passed in validation_results:
        status = "✓ PASS" if passed else "✗ FAIL"
        print(f"  {test_name:.<40} {status}")
        if not passed:
            all_passed = False
    
    print("\n" + "="*80)
    if all_passed:
        print("  ✓ ALL VALIDATIONS PASSED")
    else:
        print("  ✗ SOME VALIDATIONS FAILED")
    print("="*80 + "\n")
    
    return all_passed

if __name__ == '__main__':
    # Check if Excel file path is provided
    excel_file = None
    if len(sys.argv) > 1:
        excel_file = sys.argv[1]
        print(f"\nValidating with input Excel: {excel_file}")
    
    success = run_complete_validation(excel_file)
    
    sys.exit(0 if success else 1)
