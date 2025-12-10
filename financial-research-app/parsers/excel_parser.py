# Excel Parser
# Robust Excel file processor for financial data templates

from pathlib import Path
from typing import List, Dict, Optional, Tuple
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from config.logging_config import get_logger

logger = get_logger('excel_parser')

class ExcelParser:
    """Excel file parser for financial research templates"""
    
    def __init__(self, file_path: str):
        """
        Initialize parser with Excel file
        
        Args:
            file_path: Path to Excel file
        """
        self.file_path = Path(file_path)
        self.workbook = None
        self.validation_errors = []
        
    def load_file(self) -> bool:
        """
        Load Excel file
        
        Returns:
            True if successful, False otherwise
        """
        try:
            if not self.file_path.exists():
                self.validation_errors.append(f"File not found: {self.file_path}")
                return False
            
            self.workbook = openpyxl.load_workbook(self.file_path, data_only=True)
            logger.info(f"Loaded Excel file: {self.file_path}")
            return True
            
        except Exception as e:
            self.validation_errors.append(f"Error loading file: {e}")
            logger.error(f"Failed to load Excel: {e}")
            return False
    
    def get_sheet_names(self) -> List[str]:
        """Get all sheet names in workbook"""
        if self.workbook:
            return self.workbook.sheetnames
        return []
    
    def extract_companies_from_sheet(self, sheet_name: str = None, 
                                    start_row: int = 2, 
                                    company_col: int = 1) -> List[str]:
        """
        Extract company names from a sheet
        
        Args:
            sheet_name: Name of sheet (uses first sheet if None)
            start_row: Row to start reading from (1-indexed)
            company_col: Column containing company names (1-indexed)
            
        Returns:
            List of company names
        """
        if not self.workbook:
            logger.error("Workbook not loaded")
            return []
        
        try:
            sheet = self.workbook[sheet_name] if sheet_name else self.workbook.active
            companies = []
            
            # Read until we hit an empty cell
            row = start_row
            while True:
                cell_value = sheet.cell(row=row, column=company_col).value
                
                if cell_value is None or str(cell_value).strip() == '':
                    break
                
                company_name = str(cell_value).strip()
                if company_name:
                    companies.append(company_name)
                
                row += 1
            
            logger.info(f"Extracted {len(companies)} companies from {sheet_name or 'active sheet'}")
            return companies
            
        except Exception as e:
            self.validation_errors.append(f"Error extracting companies: {e}")
            logger.error(f"Error extracting companies: {e}")
            return []
    
    def find_header_row(self, sheet: Worksheet, search_terms: List[str]) -> Optional[int]:
        """
        Find row containing header terms
        
        Args:
            sheet: Worksheet to search
            search_terms: Terms to look for in header
            
        Returns:
            Row number (1-indexed) or None
        """
        for row in range(1, min(20, sheet.max_row + 1)):  # Search first 20 rows
            row_values = [str(cell.value).lower() if cell.value else '' 
                         for cell in sheet[row]]
            
            if any(term.lower() in ' '.join(row_values) for term in search_terms):
                return row
        
        return None
    
    def validate_template_structure(self, expected_sheets: List[str] = None) -> Tuple[bool, List[str]]:
        """
        Validate Excel template structure
        
        Args:
            expected_sheets: List of expected sheet names
            
        Returns:
            Tuple of (is_valid, errors)
        """
        errors = []
        
        if not self.workbook:
            errors.append("Workbook not loaded")
            return False, errors
        
        sheet_names = self.get_sheet_names()
        
        if not sheet_names:
            errors.append("No sheets found in workbook")
            return False, errors
        
        if expected_sheets:
            for expected in expected_sheets:
                if expected not in sheet_names:
                    errors.append(f"Expected sheet '{expected}' not found")
        
        is_valid = len(errors) == 0
        
        if is_valid:
            logger.info("Template structure validation passed")
        else:
            logger.warning(f"Template validation failed: {len(errors)} errors")
        
        return is_valid, errors
    
    def extract_quarterly_data(self, sheet_name: str, 
                              company_row: int,
                              quarter_cols: Dict[str, int]) -> Dict[str, float]:
        """
        Extract quarterly financial data for a company
        
        Args:
            sheet_name: Sheet containing data
            company_row: Row containing company data
            quarter_cols: Dictionary mapping quarter names to column indices
            
        Returns:
            Dictionary of quarter: value pairs
        """
        if not self.workbook:
            return {}
        
        try:
            sheet = self.workbook[sheet_name]
            quarterly_data = {}
            
            for quarter, col in quarter_cols.items():
                value = sheet.cell(row=company_row, column=col).value
                if value is not None and isinstance(value, (int, float)):
                    quarterly_data[quarter] = float(value)
            
            return quarterly_data
            
        except Exception as e:
            logger.error(f"Error extracting quarterly data: {e}")
            return {}
    
    def get_validation_errors(self) -> List[str]:
        """Get all validation errors"""
        return self.validation_errors
    
    def close(self):
        """Close the workbook"""
        if self.workbook:
            self.workbook.close()
            logger.info("Workbook closed")


def create_sample_template(output_path: str, companies: List[str]):
    """
    Create a sample Excel template for testing
    
    Args:
        output_path: Path to save template
        companies: List of companies to include
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Quarterly Financials"
    
    # Headers
    ws['A1'] = 'Company'
    ws['B1'] = 'Q1 FY24'
    ws['C1'] = 'Q2 FY24'
    ws['D1'] = 'Q3 FY24'
    ws['E1'] = 'Q4 FY24'
    
    # Companies
    for idx, company in enumerate(companies, start=2):
        ws[f'A{idx}'] = company
    
    wb.save(output_path)
    logger.info(f"Created sample template: {output_path}")
