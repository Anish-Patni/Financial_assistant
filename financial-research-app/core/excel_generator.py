# Excel Generator
# Create formatted Excel workbooks with financial research data

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path
from typing import Dict, List, Optional
import re
from config.logging_config import get_logger

logger = get_logger('excel_generator')

class ExcelGenerator:
    """Generate formatted Excel workbooks from research data"""
    
    def __init__(self):
        """Initialize Excel generator"""
        self.workbook = None
        
    def create_workbook(self, research_results: List[Dict], template_path: Optional[str] = None) -> Workbook:
        """
        Create Excel workbook from research results
        
        Args:
            research_results: List of research result dictionaries
            template_path: Optional path to template Excel file
            
        Returns:
            Workbook object
        """
        if template_path and Path(template_path).exists():
            try:
                self.workbook = load_workbook(template_path)
                self._fill_workbook_from_template(research_results)
                return self.workbook
            except Exception as e:
                logger.error(f"Failed to load template {template_path}: {e}")
                # Fallback to creating new
                self.workbook = Workbook()
        else:
            self.workbook = Workbook()
        
        # Default creation logic (fallback or new)
        if 'Sheet' in self.workbook.sheetnames:
            del self.workbook['Sheet']
        
        # Create Consolidated Dashboard (First Sheet)
        self._create_consolidated_dashboard(research_results)
        
        self._create_summary_sheet(research_results)
        
        # Group by company to avoid dupes if filling generic sheets
        companies = set(r.get('company') for r in research_results)
        if not template_path:
             for result in research_results:
                self._create_company_sheet(result)
        
        logger.info(f"Created workbook with {len(self.workbook.sheetnames)} sheets")
        return self.workbook

    def _create_consolidated_dashboard(self, results: List[Dict]):
        """Create a consolidated dashboard sheet with all companies"""
        ws = self.workbook.create_sheet("Consolidated Dashboard", 0)
        
        # Title
        ws['A1'] = "Consolidated Financial Dashboard"
        ws['A1'].font = Font(size=18, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="2c3e50", end_color="2c3e50", fill_type="solid")
        ws.merge_cells('A1:L1')
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 35
        
        # Headers
        headers = [
            'Company', 'Quarter', 'Year', 
            'Revenue', 'EBITDA', 'EBITDA %', 
            'EBIT', 'EBIT %', 'PBT', 'PAT', 
            'EPS', 'Status'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="34495e", end_color="34495e", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
            
            # Set column width
            width = 15
            if header == 'Company': width = 25
            elif 'Revenue' in header or 'EBITDA' in header: width = 18
            ws.column_dimensions[get_column_letter(col)].width = width

        # Sort results by Company, then Year, then Quarter
        sorted_results = sorted(results, key=lambda x: (x.get('company', ''), x.get('year', 0), x.get('quarter', '')))
        
        # Data Rows
        row = 4
        for res in sorted_results:
            data = res.get('extracted_data', {})
            
            # Helper to get value safely
            def get_val(key):
                if key in data:
                    return data[key].get('value')
                return None

            # 1. Company
            ws.cell(row=row, column=1, value=res.get('company', 'N/A'))
            
            # 2. Quarter
            ws.cell(row=row, column=2, value=res.get('quarter', 'N/A'))
            ws.cell(row=row, column=2).alignment = Alignment(horizontal='center')
            
            # 3. Year
            ws.cell(row=row, column=3, value=res.get('year', 'N/A'))
            ws.cell(row=row, column=3).alignment = Alignment(horizontal='center')
            
            # 4. Revenue
            rev = get_val('total_income')
            if rev:
                c = ws.cell(row=row, column=4, value=rev)
                c.number_format = '#,##0.00'
            else:
                ws.cell(row=row, column=4, value='--').alignment = Alignment(horizontal='center')

            # 5. EBITDA
            ebitda = get_val('ebitda')
            if ebitda:
                c = ws.cell(row=row, column=5, value=ebitda)
                c.number_format = '#,##0.00'
            else:
                ws.cell(row=row, column=5, value='--').alignment = Alignment(horizontal='center')

            # 6. EBITDA %
            ebitda_margin = get_val('ebitda_margin')
            if ebitda_margin:
                c = ws.cell(row=row, column=6, value=ebitda_margin/100 if ebitda_margin > 1 else ebitda_margin) # Handle both 15.5 and 0.155
                c.number_format = '0.00%'
            elif ebitda and rev:
                try:
                    val = ebitda / rev
                    c = ws.cell(row=row, column=6, value=val)
                    c.number_format = '0.00%'
                except:
                    ws.cell(row=row, column=6, value='--')
            else:
                ws.cell(row=row, column=6, value='--').alignment = Alignment(horizontal='center')

            # 7. EBIT
            ebit = get_val('ebit')
            if ebit:
                c = ws.cell(row=row, column=7, value=ebit)
                c.number_format = '#,##0.00'
            else:
                ws.cell(row=row, column=7, value='--').alignment = Alignment(horizontal='center')

            # 8. EBIT %
            ebit_margin = get_val('ebit_margin')
            if ebit_margin:
                c = ws.cell(row=row, column=8, value=ebit_margin/100 if ebit_margin > 1 else ebit_margin)
                c.number_format = '0.00%'
            elif ebit and rev:
                try:
                    val = ebit / rev
                    c = ws.cell(row=row, column=8, value=val)
                    c.number_format = '0.00%'
                except:
                    ws.cell(row=row, column=8, value='--')
            else:
                ws.cell(row=row, column=8, value='--').alignment = Alignment(horizontal='center')

            # 9. PBT
            pbt = get_val('pbt')
            if pbt:
                c = ws.cell(row=row, column=9, value=pbt)
                c.number_format = '#,##0.00'
            else:
                ws.cell(row=row, column=9, value='--').alignment = Alignment(horizontal='center')

            # 10. PAT
            pat = get_val('pat') or get_val('net_profit')
            if pat:
                c = ws.cell(row=row, column=10, value=pat)
                c.number_format = '#,##0.00'
            else:
                ws.cell(row=row, column=10, value='--').alignment = Alignment(horizontal='center')

            # 11. EPS
            eps = get_val('eps')
            if eps:
                c = ws.cell(row=row, column=11, value=eps)
                c.number_format = '0.00'
            else:
                ws.cell(row=row, column=11, value='--').alignment = Alignment(horizontal='center')

            # 12. Status
            status = res.get('status', 'unknown')
            c = ws.cell(row=row, column=12, value=status.upper())
            c.alignment = Alignment(horizontal='center')
            if status == 'success':
                c.font = Font(color="006400", bold=True)
            else:
                c.font = Font(color="8B0000", bold=True)

            row += 1
        
        # Add borders
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        for r in range(3, row):
            for c in range(1, 13):
                ws.cell(row=r, column=c).border = thin_border
        """Fill existing workbook using smart matching"""
        # Group results by company
        company_data = {}
        for res in results:
            company = res.get('company')
            if not company: continue
            if company not in company_data:
                company_data[company] = []
            company_data[company].append(res)

        # Identify template sheet (Sheet 3 or "Company-wise" or index 2)
        template_sheet = None
        if len(self.workbook.sheetnames) >= 3:
             template_sheet = self.workbook.worksheets[2] # Index 2 = Sheet 3
        else:
             template_sheet = self.workbook.worksheets[-1] # Last sheet

        # Process each company
        for company, company_results in company_data.items():
            # 1. Fill Summary/Sheet 1 if possible
            # Scan first sheet for company name row
            if len(self.workbook.sheetnames) > 0:
                 self._smart_fill_sheet(self.workbook.worksheets[0], company, company_results)

            # 2. Create detailed sheet from template - REMOVED to avoid duplicates
            # The _create_simple_company_sheet below creates the correct [Company]_Data sheet
            
            # 3. Always create a simple data sheet as well for visibility
            self._create_simple_company_sheet(company, company_results)

    def _smart_fill_sheet(self, ws, company: str, results: List[Dict]):
        """
        Intelligently fill a worksheet by matching headers and quarters.
        Looks for:
        1. "Revenue", "EBITDA" etc in cells -> Maps to metric
        2. "Q1", "Q2", "2024" etc in headers -> Maps to time period
        3. Fills intersection
        """
        logger.info(f"Filling sheet '{ws.title}' for company '{company}' with {len(results)} results")
        
        # Map standardized keys to likely display labels
        metric_map = {
            'total_income': ['revenue', 'total income', 'sales', 'top line'],
            'ebitda': ['ebitda', 'operating profit'],
            'ebitda_margin': ['ebitda %', 'op. ebitda%', 'ebitda margin', 'op. ebitda %'],
            'ebit': ['ebit', 'operating ebit'],
            'ebit_margin': ['ebit %', 'op. ebit%', 'ebit margin', 'op. ebit %'],
            'op_pbt': ['op. pbt', 'op pbt', 'operating pbt'],
            'pbt': ['pbt', 'profit before tax'],
            'net_profit': ['pat', 'net profit', 'profit after tax'],
            'eps': ['eps', 'earning per share'],
            'interest': ['interest', 'finance cost'],
            'other_income': ['other income', 'non-operating income']
        }

        # Index the sheet headers (first 10 rows, first 20 cols)
        rows_to_scan = 20
        cols_to_scan = 20
        
        # Locate Quarter columns
        quarter_cols = {} # { 'Q1': col_idx, 'Q2': col_idx ... }
        
        # Scan for header row containing Q1/Q2...
        header_row = None
        
        for r in range(1, rows_to_scan):
            row_values = []
            for c in range(1, cols_to_scan):
                val = ws.cell(row=r, column=c).value
                if isinstance(val, str):
                    val_lower = val.lower()
                    if 'q1' in val_lower: quarter_cols['Q1'] = c
                    elif 'q2' in val_lower: quarter_cols['Q2'] = c
                    elif 'q3' in val_lower: quarter_cols['Q3'] = c
                    elif 'q4' in val_lower: quarter_cols['Q4'] = c
            
            if len(quarter_cols) >= 1:
                header_row = r
                break
        
        if not header_row:
             logger.warning(f"No quarter headers found in sheet '{ws.title}' for {company}")
             # Try vertical layout (Quarters in simple column A or B?)
             # Skip for now, assuming horizontal time layout based on user screenshot
             return
        
        logger.info(f"Found quarter headers at row {header_row}: {quarter_cols}")

        # Iterate rows to find metrics
        for r in range(header_row + 1, rows_to_scan + 20): # Scan deeper for metrics
            # Check first few columns for metric label
            metric_found = None
            cell_val = None
            
            for c in range(1, 5): # Metric label usually in first columns
                val = ws.cell(row=r, column=c).value
                if isinstance(val, str):
                    val_lower = val.lower().strip()
                    # Check against map
                    for key, keywords in metric_map.items():
                        if any(k in val_lower for k in keywords):
                            metric_found = key
                            cell_val = val
                            break
                if metric_found: break
            
            if metric_found:
                logger.info(f"Found metric '{metric_found}' at row {r}, label: '{cell_val}'")
                # Fill data for each quarter column
                for q_label, col_idx in quarter_cols.items():
                    # Find result for this quarter
                    res = next((res for res in results if res.get('quarter') == q_label), None)
                    if res:
                        logger.debug(f"  Processing {q_label} at column {col_idx}")
                        data = res.get('extracted_data', {})
                        
                        # Get value
                        value = None
                        if metric_found in data:
                            value = data[metric_found].get('value')
                        
                        # Handle derived metrics if missing
                        if value is None:
                             if metric_found == 'ebitda_margin' and 'ebitda' in data and 'total_income' in data:
                                   try:
                                       value = (data['ebitda']['value'] / data['total_income']['value']) * 100
                                   except:
                                       pass
                             elif metric_found == 'ebit_margin' and 'ebit' in data and 'total_income' in data:
                                   try:
                                       value = (data['ebit']['value'] / data['total_income']['value']) * 100
                                   except:
                                       pass
                             # Op PBT - always calculate from EBIT - Interest
                             elif metric_found == 'op_pbt' and 'ebit' in data and 'interest' in data:
                                    try:
                                        value = data['ebit']['value'] - data['interest']['value']
                                    except:
                                        pass

                        if value is not None:
                            try:
                                # Check if cell is part of a merged range and unmerge if needed
                                from openpyxl.worksheet.cell_range import CellRange
                                cell_coordinate = ws.cell(row=r, column=col_idx).coordinate
                                
                                # Find and unmerge if this cell is in a merged range
                                for merged_range in list(ws.merged_cells.ranges):
                                    if cell_coordinate in merged_range:
                                        ws.unmerge_cells(str(merged_range))
                                        logger.debug(f"  Unmerged {merged_range} to write to {cell_coordinate}")
                                        break
                                
                                # Now write the value
                                ws.cell(row=r, column=col_idx, value=value)
                                logger.info(f"  OK Filled {metric_found}={value} at row {r}, col {col_idx} ({q_label})")
                            except Exception as e:
                                logger.error(f"  ERROR writing to row {r}, col {col_idx}: {e}")
                        else:
                            logger.debug(f"  MISS Could not get value for {metric_found} in {q_label}")


    def _create_summary_sheet(self, results: List[Dict]):
        """Create summary overview sheet"""
        ws = self.workbook.create_sheet("Summary", 0)
        
        # Title
        ws['A1'] = "Financial Research Summary"
        ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
        ws.merge_cells('A1:F1')
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 30
        
        # Headers
        headers = ['Company', 'Quarter', 'Year', 'Indicators', 'Avg Confidence', 'Status']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="764ba2", end_color="764ba2", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
        
        # Data rows
        for row_idx, result in enumerate(results, 4):
            extracted = result.get('extracted_data', {})
            avg_conf = sum(v['confidence'] for v in extracted.values()) / len(extracted) if extracted else 0
            
            ws.cell(row=row_idx, column=1, value=result.get('company', 'N/A'))
            ws.cell(row=row_idx, column=2, value=result.get('quarter', 'N/A'))
            ws.cell(row=row_idx, column=3, value=result.get('year', 'N/A'))
            ws.cell(row=row_idx, column=4, value=len(extracted))
            
            conf_cell = ws.cell(row=row_idx, column=5, value=f"{avg_conf*100:.1f}%")
            # Color code confidence
            if avg_conf >= 0.9:
                conf_cell.fill = PatternFill(start_color="d4edda", end_color="d4edda", fill_type="solid")
            elif avg_conf >= 0.7:
                conf_cell.fill = PatternFill(start_color="fff3cd", end_color="fff3cd", fill_type="solid")
            else:
                conf_cell.fill = PatternFill(start_color="f8d7da", end_color="f8d7da", fill_type="solid")
            
            status_cell = ws.cell(row=row_idx, column=6, value=result.get('status', 'unknown'))
            if result.get('status') == 'success':
                status_cell.fill = PatternFill(start_color="d4edda", end_color="d4edda", fill_type="solid")
        
        # Auto-adjust column widths
        for col in range(1, 7):
            ws.column_dimensions[get_column_letter(col)].width = 18
    
    def _create_company_sheet(self, result: Dict):
        """Create detailed sheet for individual company"""
        company = result.get('company', 'Unknown')
        quarter = result.get('quarter', 'Q1')
        year = result.get('year', 2024)
        
        sheet_name = f"{company[:25]}_{quarter}_{year}"  # Truncate if too long
        ws = self.workbook.create_sheet(sheet_name)
        
        # Header
        ws['A1'] = f"{company} - {quarter} {year}"
        ws['A1'].font = Font(size=14, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
        ws.merge_cells('A1:D1')
        ws['A1'].alignment = Alignment(horizontal='center')
        ws.row_dimensions[1].height = 25
        
        # Context confidence
        context_conf = result.get('context_confidence', 0)
        ws['A3'] = "Context Confidence:"
        ws['B3'] = f"{context_conf*100:.1f}%"
        ws['A3'].font = Font(bold=True)
        
        # Extracted indicators header
        ws['A5'] = "Financial Indicators"
        ws['A5'].font = Font(size=12, bold=True)
        ws.merge_cells('A5:D5')
        
        # Table headers
        indicator_headers = ['Indicator', 'Value (₹ Cr)', 'Confidence', 'Status']
        for col, header in enumerate(indicator_headers, 1):
            cell = ws.cell(row=6, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="764ba2", end_color="764ba2", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
        
        # Indicator data
        extracted = result.get('extracted_data', {})
        row = 7
        for indicator, data in extracted.items():
            ws.cell(row=row, column=1, value=indicator.replace('_', ' ').title())
            ws.cell(row=row, column=2, value=f"₹{data['value']:,.2f}")
            
            conf = data['confidence']
            conf_cell = ws.cell(row=row, column=3, value=f"{conf*100:.0f}%")
            
            # Color code confidence
            if conf >= 0.9:
                conf_cell.fill = PatternFill(start_color="d4edda", end_color="d4edda", fill_type="solid")
                ws.cell(row=row, column=4, value="✓ High")
            elif conf >= 0.7:
                conf_cell.fill = PatternFill(start_color="fff3cd", end_color="fff3cd", fill_type="solid")
                ws.cell(row=row, column=4, value="⚠ Medium")
            else:
                conf_cell.fill = PatternFill(start_color="f8d7da", end_color="f8d7da", fill_type="solid")
                ws.cell(row=row, column=4, value="⚠ Low")
            
            row += 1
        
        # Add borders
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row_idx in range(6, row):
            for col_idx in range(1, 5):
                ws.cell(row=row_idx, column=col_idx).border = thin_border
        
        # Auto-adjust columns
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
    
    def _create_simple_company_sheet(self, company: str, results: List[Dict]):
        """Create a simple, readable sheet with all financial data for a company"""
        sheet_name = f"{company[:25]}_Data"
        try:
            ws = self.workbook.create_sheet(sheet_name)
        except:
            # Sheet name might already exist, add suffix
            ws = self.workbook.create_sheet(f"{sheet_name}_{len(self.workbook.sheetnames)}")
        
        # Title
        ws['A1'] = f"{company} - Financial Data"
        ws['A1'].font = Font(size=14, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
        ws.merge_cells('A1:F1')
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 25
        
        # Headers
        ws['A3'] = "Metric"
        ws['A3'].font = Font(bold=True)
        
        # Sort results by quarter
        sorted_results = sorted(results, key=lambda x: (x.get('year', 0), x.get('quarter', 'Q1')))
        
        col_idx = 2
        quarter_cols = {}
        for result in sorted_results:
            q_label = f"{result.get('quarter', 'Q?')} {result.get('year', '?')}"
            ws.cell(row=3, column=col_idx, value=q_label)
            ws.cell(row=3, column=col_idx).font = Font(bold=True)
            ws.cell(row=3, column=col_idx).alignment = Alignment(horizontal='center')
            quarter_cols[result.get('quarter')] = col_idx
            col_idx += 1
        
        # Metrics to display
        metrics_display = [
            ('Revenue', 'total_income', False),
            ('Op. EBITDA%', 'ebitda_margin', True),
            ('Op. EBIT%', 'ebit_margin', True),
            ('Op. PBT', 'op_pbt', False),
            ('PBT', 'pbt', False),
            ('PAT', 'pat', False),
            ('EBITDA', 'ebitda', False),
            ('EBIT', 'ebit', False),
            ('Interest', 'interest', False),
            ('Other Income', 'other_income', False),
        ]
        
        row = 4
        for label, key, is_percent in metrics_display:
            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=1).font = Font(bold=True)
            
            col_idx = 2
            for result in sorted_results:
                data = result.get('extracted_data', {})
                value = None
                
                # Get or calculate value
                if key in data:
                    value = data[key].get('value')
                elif key == 'op_pbt' and 'ebit' in data and 'interest' in data:
                    # Calculate Op. PBT
                    try:
                        value = data['ebit']['value'] - data['interest']['value']
                    except:
                        pass
                elif key == 'ebitda_margin' and 'ebitda' in data and 'total_income' in data:
                    # Calculate EBITDA margin
                    try:
                        value = (data['ebitda']['value'] / data['total_income']['value']) * 100
                    except:
                        pass
                elif key == 'ebit_margin' and 'ebit' in data and 'total_income' in data:
                    # Calculate EBIT margin
                    try:
                        value = (data['ebit']['value'] / data['total_income']['value']) * 100
                    except:
                        pass
                
                if value is not None:
                    if is_percent:
                        ws.cell(row=row, column=col_idx, value=f"{value:.2f}%")
                    else:
                        ws.cell(row=row, column=col_idx, value=value)
                        ws.cell(row=row, column=col_idx).number_format = '#,##0.00'
                else:
                    ws.cell(row=row, column=col_idx, value='--')
                
                ws.cell(row=row, column=col_idx).alignment = Alignment(horizontal='right')
                col_idx += 1
            
            row += 1
        
        # Auto-adjust columns
        ws.column_dimensions['A'].width = 20
        for col_idx in range(2, 2 + len(sorted_results)):
            ws.column_dimensions[get_column_letter(col_idx)].width = 15
        
        logger.info(f"Created simple data sheet '{ws.title}' for {company}")
    
    def save_workbook(self, filepath: Path) -> bool:
        """
        Save workbook to file
        
        Args:
            filepath: Path to save file
            
        Returns:
            True if successful
        """
        try:
            filepath.parent.mkdir(parents=True, exist_ok=True)
            self.workbook.save(filepath)
            logger.info(f"Saved workbook to {filepath}")
            return True
        except Exception as e:
            logger.error(f"Error saving workbook: {e}")
            return False
    
    def generate_excel(self, research_results: List[Dict], output_path: Path, template_path: Optional[str] = None) -> bool:
        """
        Generate and save Excel file in one step
        
        Args:
            research_results: List of research results
            output_path: Path to save file
            template_path: Optional path to template
            
        Returns:
            True if successful
        """
        try:
            self.create_workbook(research_results, template_path)
            return self.save_workbook(output_path)
        except Exception as e:
            logger.error(f"Error generating Excel: {e}")
            return False
