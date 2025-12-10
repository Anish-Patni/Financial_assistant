#!/usr/bin/env python3
"""Check the generated Excel file"""

from openpyxl import load_workbook

wb = load_workbook('data/output/test_financial_research.xlsx')

print("=" * 80)
print(f"Generated Excel Sheets: {wb.sheetnames}")
print("=" * 80)

# Check one of the data sheets
if 'Persistent_Data' in wb.sheetnames:
    ws = wb['Persistent_Data']
    print(f"\nPersistent_Data sheet content:")
    print("-" * 80)
    for row_idx in range(1, 15):
        row_data = []
        for col_idx in range(1, 5):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val:
                row_data.append(str(val))
        if row_data:
            print(f"Row {row_idx:2d}: {' | '.join(row_data)}")
    print("=" * 80)

print("\n✓ Excel file generated successfully with data sheets!")
print(f"  Location: data/output/test_financial_research.xlsx")
print(f"  Total Sheets: {len(wb.sheetnames)}")
