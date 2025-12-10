"""
Excel Upload Handler
Processes uploaded Excel files and extracts company lists
"""

from pathlib import Path
from typing import List, Dict, Optional, Tuple
import openpyxl
from openpyxl import load_workbook
from config.logging_config import get_logger

logger = get_logger('excel_upload_handler')


class ExcelUploadHandler:
    """Handle Excel file uploads and company extraction"""
    
    def __init__(self):
        self.logger = logger
        self.supported_extensions = ['.xlsx', '.xls']
        
    def validate_file(self, filepath: Path) -> Tuple[bool, Optional[str]]:
        """
        Validate uploaded Excel file
        
        Args:
            filepath: Path to uploaded file
            
        Returns:
            Tuple of (is_valid, error_message)
        """
        # Check file exists
        if not filepath.exists():
            return False, "File not found"
        
        # Check extension
        if filepath.suffix.lower() not in self.supported_extensions:
            return False, f"Invalid file type. Supported: {', '.join(self.supported_extensions)}"
        
        # Check file size (max 10MB)
        max_size = 10 * 1024 * 1024  # 10MB
        if filepath.stat().st_size > max_size:
            return False, f"File too large. Maximum size is 10MB"
        
        # Try to open file
        try:
            wb = load_workbook(filepath, read_only=True)
            wb.close()
        except Exception as e:
            return False, f"Invalid Excel file: {str(e)}"
        
        return True, None
    
    def extract_companies(self, filepath: Path, sheet_name: Optional[str] = None) -> List[str]:
        """
        Extract company names from Excel file
        
        Args:
            filepath: Path to Excel file
            sheet_name: Optional sheet name (uses first sheet if not specified)
            
        Returns:
            List of company names
        """
        try:
            wb = load_workbook(filepath, read_only=True, data_only=True)
            
            # Get sheet
            if sheet_name and sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
            else:
                # Use first sheet or look for common sheet names
                common_names = ['Quarterly Financials', 'Companies', 'Sheet1', 'Data']
                sheet = None
                for name in common_names:
                    if name in wb.sheetnames:
                        sheet = wb[name]
                        break
                
                if sheet is None:
                    sheet = wb.active
            
            self.logger.info(f"Extracting companies from sheet: {sheet.title}")
            
            # Extract companies
            companies = self._extract_companies_from_sheet(sheet)
            
            wb.close()
            
            self.logger.info(f"Extracted {len(companies)} companies: {companies}")
            return companies
            
        except Exception as e:
            self.logger.error(f"Error extracting companies: {e}")
            raise
    
    def _extract_companies_from_sheet(self, sheet) -> List[str]:
        """
        Extract companies from a worksheet
        Looks for a column with header 'Companies' or similar
        
        Args:
            sheet: Worksheet object
            
        Returns:
            List of company names
        """
        companies = []
        
        # Strategy 1: Look for 'Companies' column header
        company_col = None
        for col_idx in range(1, min(sheet.max_column + 1, 10)):  # Check first 10 columns
            cell_value = sheet.cell(row=1, column=col_idx).value
            if cell_value and isinstance(cell_value, str):
                if cell_value.strip().lower() in ['companies', 'company', 'company name', 'firm']:
                    company_col = col_idx
                    break
        
        if company_col:
            # Extract from identified column
            for row_idx in range(2, sheet.max_row + 1):
                cell_value = sheet.cell(row=row_idx, column=company_col).value
                if cell_value and isinstance(cell_value, str):
                    company_name = cell_value.strip()
                    if company_name and company_name not in companies:
                        companies.append(company_name)
        else:
            # Strategy 2: Extract from first column (A column)
            self.logger.info("No 'Companies' header found, using first column")
            for row_idx in range(2, min(sheet.max_row + 1, 100)):  # Check first 100 rows
                cell_value = sheet.cell(row=row_idx, column=1).value
                if cell_value and isinstance(cell_value, str):
                    company_name = cell_value.strip()
                    # Filter out empty strings and common non-company values
                    if (company_name and 
                        company_name not in companies and
                        not company_name.lower() in ['total', 'average', 'sum', 'revenue growth', 'op. ebitda growth', 'growth']):
                        companies.append(company_name)
        
        # Clean up company names
        companies = [c for c in companies if c and len(c) > 1]
        
        return companies
    
    def get_file_info(self, filepath: Path) -> Dict:
        """
        Get information about uploaded Excel file
        
        Args:
            filepath: Path to file
            
        Returns:
            Dictionary with file information
        """
        try:
            wb = load_workbook(filepath, read_only=True)
            
            info = {
                'filename': filepath.name,
                'size_bytes': filepath.stat().st_size,
                'sheet_count': len(wb.sheetnames),
                'sheet_names': wb.sheetnames
            }
            
            # Get row/column info from active sheet
            sheet = wb.active
            info['rows'] = sheet.max_row
            info['columns'] = sheet.max_column
            
            wb.close()
            
            return info
        except Exception as e:
            self.logger.error(f"Error getting file info: {e}")
            return {
                'filename': filepath.name,
                'error': str(e)
            }
    
    def preview_data(self, filepath: Path, max_rows: int = 10) -> Dict:
        """
        Get preview of Excel data
        
        Args:
            filepath: Path to file
            max_rows: Maximum rows to preview
            
        Returns:
            Dictionary with preview data
        """
        try:
            wb = load_workbook(filepath, read_only=True, data_only=True)
            sheet = wb.active
            
            preview = {
                'sheet_name': sheet.title,
                'headers': [],
                'rows': []
            }
            
            # Get headers
            for col_idx in range(1, min(sheet.max_column + 1, 20)):
                cell_value = sheet.cell(row=1, column=col_idx).value
                preview['headers'].append(str(cell_value) if cell_value else '')
            
            # Get data rows
            for row_idx in range(2, min(max_rows + 2, sheet.max_row + 1)):
                row_data = []
                for col_idx in range(1, min(sheet.max_column + 1, 20)):
                    cell_value = sheet.cell(row=row_idx, column=col_idx).value
                    row_data.append(str(cell_value) if cell_value else '')
                preview['rows'].append(row_data)
            
            wb.close()
            
            return preview
        except Exception as e:
            self.logger.error(f"Error previewing data: {e}")
            return {
                'error': str(e)
            }
